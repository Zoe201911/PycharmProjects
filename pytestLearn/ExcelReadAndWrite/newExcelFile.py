#author:Zoe
#createTime:2019-12-17
#读取一个Excel文件

import openpyxl
def create_Excel():
    f = openpyxl.Workbook()
    print(f.sheetnames)
    sheet1 = f.active
    sheet1.title = 'hello'
    sheet1 = f.create_sheet(title='kitty')
    print(f.sheetnames)
    rows_num = 7
    cols_num = 4
    for rown in range(rows_num):
        for coln in range(cols_num):
            rw = rown + 1
            cl = coln + 1
            if coln%2 ==0:
                sheet1.cell(row = rw,column = cl,value='1')
            else:
                sheet1.cell(row = rw,column = cl ,value='2')
    f.save('file/Demo.xlsx')

if __name__ == '__main__':
    create_Excel()


